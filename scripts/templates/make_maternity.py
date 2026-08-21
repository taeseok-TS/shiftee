# 출산휴가원 — 엑셀에 필드를 심은 뒤 워드로 변환
#
# 순서: ① openpyxl 로 셀에 {필드}·《마커》 기입 → ② Excel COM 복사 →
#       Word PasteSpecial(HTML) 로 편집 가능한 워드 표 → ③ docx 저장
# (엔진은 워드 전용이라 엑셀은 필드채움·서명이 안 된다 — 사직원 때 확립한 패턴)
#
# 결재칸 매핑: 담당=근로자(신청자), 검토=원장, 승인=본부
import os, shutil
import openpyxl

SRC = r"C:/Users/N-88/Downloads/출산휴가원 양식.xlsx"
MID = os.path.abspath("출산휴가원_필드.xlsx")
OUT = os.path.abspath("출산휴가원_템플릿.docx")

# ── 1) 엑셀에 필드 기입 ─────────────────────────────────────
shutil.copy(SRC, MID)
wb = openpyxl.load_workbook(MID)
ws = wb.active

ws["H3"] = "《근로자서명》"   # 담당(신청자)
ws["I3"] = "《원장서명》"     # 검토
ws["J3"] = "《본부서명》"     # 승인

ws["C5"] = "{직원명}"
ws["C6"] = "{지점}"
ws["H6"] = "{직급}"

ws["E8"] = "{휴가시작일}"
ws["E9"] = "{휴가종료일}"
ws["D10"] = "출산 (출산 예정일 : {출산예정일})"

ws["D11"] = "{주소}"
ws["D12"] = "{연락처}"
ws["D13"] = "{생년월일}"
ws["C14"] = "{부재중 연락처}"

# 신청일·신청인 — 좁은 셀에서 줄바꿈되지 않게 병합해 한 칸으로
from openpyxl.styles import Alignment
ws["G16"] = "신청일 : {작성일}"
ws["H16"] = ""; ws["I16"] = ""; ws["J16"] = ""
ws.merge_cells("G16:J16")
ws["G16"].alignment = Alignment(horizontal="right")

ws["H18"] = "신청인 : {직원명}   《근로자서명》   (인)"
ws["I18"] = ""
ws.merge_cells("H18:J18")
ws["H18"].alignment = Alignment(horizontal="right")

# 페이지 넘침·꺾임 방지 — 과대한 행 높이 축소, 제목 폰트 축소, 좁은 G열 확대
from openpyxl.styles import Font
t_font = ws["B2"].font
ws["B2"].font = Font(name=t_font.name, size=26, bold=t_font.bold)
ws.column_dimensions["G"].width = 12
for r, h in {3: 42, 5: 30, 6: 30, 7: 24, 8: 26, 9: 26, 10: 34, 11: 28, 12: 28, 13: 28, 14: 36}.items():
    ws.row_dimensions[r].height = h

# 노란 배경 제거 (완성본에 남는다)
from openpyxl.styles import PatternFill
for row in ws.iter_rows():
    for c in row:
        f = c.fill
        if f and f.fgColor and str(f.fgColor.rgb or "").upper().endswith(("FFFF00", "FFFF99")):
            c.fill = PatternFill(fill_type=None)

wb.save(MID)
print("① 필드 기입 완료:", MID)

# ── 2) Excel → Word 변환 (COM) ──────────────────────────────
import win32com.client

excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False
word = win32com.client.Dispatch("Word.Application")
word.Visible = False
try:
    xwb = excel.Workbooks.Open(MID, ReadOnly=True)
    xws = xwb.Worksheets(1)
    xws.UsedRange.Copy()

    doc = word.Documents.Add()
    # 여백을 좁혀 A4 한 장에 들어가게 (1cm = 28.35pt — COM 헬퍼가 오류를 내서 직접 계산)
    doc.PageSetup.TopMargin = 1.5 * 28.35
    doc.PageSetup.BottomMargin = 1.5 * 28.35
    doc.PageSetup.LeftMargin = 1.8 * 28.35
    doc.PageSetup.RightMargin = 1.8 * 28.35
    word.Selection.PasteSpecial(DataType=10)  # 10 = HTML (편집 가능한 워드 표)

    # 표를 페이지 폭에 맞춤
    for t in doc.Tables:
        t.AutoFitBehavior(2)  # wdAutoFitWindow

    doc.SaveAs(OUT, FileFormat=16)  # 16 = docx
    pages = doc.ComputeStatistics(2)
    doc.Close(False)
    xwb.Close(False)
    print(f"② 변환 완료: {OUT} ({pages}쪽)")
finally:
    excel.CutCopyMode = False
    excel.Quit()
    word.Quit()
